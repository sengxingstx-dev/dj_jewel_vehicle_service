from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db import transaction
from django.http import Http404, HttpResponse
from django.template.loader import get_template
from django.utils import timezone
from openpyxl import Workbook
from weasyprint import HTML, CSS
import csv

from apps.account.forms import RegistrationForm
from .forms import CarForm, CategoryForm, RentalForm, PaymentForm
from .models import RentalAgent, Customer, Car, Category, Rental, Payment


#
# NOTE: Customer
#
def home_view(request):
    if request.user.is_authenticated:
        if request.user.is_superuser or request.user.is_staff:
            return redirect('dashboard')

    cars = Car.objects.all()

    context = {'cars': cars}
    return render(request, 'jewel_vehicle_service/home.html', context)


def rental_view(request, pk):
    try:
        agent = RentalAgent.objects.get(user=request.user)
    except RentalAgent.DoesNotExist:
        agent = None
    try:
        customer = Customer.objects.get(user=request.user)
    except Customer.DoesNotExist:
        customer = None

    car = Car.objects.get(id=pk)

    context = {'agent': agent, 'customer': customer, 'car': car}
    return render(request, 'jewel_vehicle_service/pages/rental-form.html', context)


@transaction.atomic
def book_now_view(request):
    if request.method == 'POST':
        # Assuming you have forms for Rental and Payment data
        rental_form = RentalForm(request.POST)
        payment_form = PaymentForm(request.POST, request.FILES)

        if rental_form.is_valid() and payment_form.is_valid():
            try:
                with transaction.atomic():
                    # Step 1: Save the Rental instance
                    rental = rental_form.save()

                    # Step 2: Use the saved Rental instance to create a new Payment instance
                    payment = payment_form.save(commit=False)
                    payment.rental = rental  # Link the Payment to the Rental
                    payment.save()  # Step 3: Save the Payment instance

            except Exception as e:
                # Handle any exceptions that might occur during the transaction
                print(f"Error: {e}")
                return redirect('error-page')  # Redirect to an error page

            return redirect('payment-receipt', pk=payment.id)  # Redirect to a success page
        else:
            # If either form is not valid, you may want to handle this case
            # For example, you can display an error message or redirect to an error page
            # Print form errors for debugging
            print("Forms are not valid.")
            print("Rental Form Errors:", rental_form.errors)
            print("Payment Form Errors:", payment_form.errors)
            return redirect('error-page')
    else:
        rental_form = RentalForm()
        payment_form = PaymentForm()

    return redirect('home')


def payment_receipt_view(request, pk):
    try:
        rental = Rental.objects.get(id=pk)
        payment = Payment.objects.get(rental=rental)
    except Payment.DoesNotExist:
        raise Http404('Payment does not exist.')

    context = {'payment': payment}
    return render(request, 'jewel_vehicle_service/pages/payment-receipt.html', context)


def export_payment_to_pdf(request, pk):
    try:
        payment = Payment.objects.get(id=pk)
    except Payment.DoesNotExist:
        raise Http404('Payment does not exist.')

    template_path = 'jewel_vehicle_service/pages/payment-receipt.html'
    context = {'payment': payment}

    # Render the template
    template = get_template(template_path)
    html_content = template.render(context)

    # Set a dynamic filename based on rental information
    filename = f'payment_receipt_{payment.id}.pdf'

    # Create a WeasyPrint HTML object
    pdf_file = HTML(string=html_content).write_pdf(stylesheets=[CSS(string='@page { size: A4; }')])

    # Create a Django response with the PDF file
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'filename="{filename}"'

    return response


def about_view(request):
    return render(request, 'jewel_vehicle_service/pages/about.html')


def service_view(request):
    return render(request, 'jewel_vehicle_service/pages/service.html')


def contact_view(request):
    return render(request, 'jewel_vehicle_service/pages/contact.html')


def portfolio_view(request):
    return render(request, 'jewel_vehicle_service/pages/portfolio.html')


def error_page_view(request):
    return render(request, 'jewel_vehicle_service/pages/error-page.html')


#
# NOTE: Dashboard
#
def dashboard_view(request):
    if request.user.is_authenticated:
        user = request.user
        agent = RentalAgent.objects.get(user=user)
        # Retrieve the latest 10 bookings
        recent_bookings = Rental.objects.order_by('-created_at')[:10]

        context = {
            'agent': agent,
            'total_agents': RentalAgent.total_agents(),
            'total_customers': Customer.total_customers(),
            'total_cars': Car.total_cars(),
            'available_cars': Car.available_cars(),
            'total_revenue': Rental.total_revenue(),
            'total_bookings': Rental.total_bookings(),
            'recent_bookings': recent_bookings,
        }
        return render(request, 'dashboard/home.html', context)
    return redirect('home')


def manage_agent_view(request):
    agent = RentalAgent.objects.get(user=request.user)

    # Get the search query and search_by from the request's GET parameters
    search_query = request.GET.get('search', '')
    search_by = request.GET.get('search_by', 'username')  # Default to searching by username

    # Customize the query based on the search_by value
    if search_query and search_by == 'username':
        agents = RentalAgent.objects.filter(user__username__icontains=search_query)
    elif search_query and search_by == 'email':
        agents = RentalAgent.objects.filter(email__icontains=search_query)
    else:
        agents = RentalAgent.objects.all().order_by('-updated_at')

    if request.method == 'POST':
        try:
            form = RegistrationForm(request.POST)
            if form.is_valid():
                user = form.save(commit=False)
                email = form.cleaned_data['email'].lower()
                first_name = form.cleaned_data.get('first_name')
                last_name = form.cleaned_data.get('last_name')
                phone = form.cleaned_data.get('phone')
                user.is_staff = True
                user.save()
                agent = RentalAgent.objects.create(
                    user=user, email=email, first_name=first_name, last_name=last_name, phone=phone
                )
                messages.success(request, 'Agent created successfully.')
            else:
                print('Form is not valid: ', form.errors)
        except Exception as e:
            print('ERROR:', str(e))

    context = {
        'agent': agent,
        'agents': agents,
    }
    return render(request, 'dashboard/html/manage-agent.html', context)


def manage_customer_view(request):
    agent = RentalAgent.objects.get(user=request.user)

    # Get the search query and search_by from the request's GET parameters
    search_query = request.GET.get('search', '')
    search_by = request.GET.get('search_by', 'username')  # Default to searching by username

    # Customize the query based on the search_by value
    if search_query and search_by == 'username':
        customers = Customer.objects.filter(user__username__icontains=search_query)
    elif search_query and search_by == 'email':
        customers = Customer.objects.filter(email__icontains=search_query)
    else:
        customers = Customer.objects.all().order_by('-updated_at')

    if request.method == 'POST':
        form = RegistrationForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            email = form.cleaned_data['email'].lower()
            user.is_staff = True
            user.save()
            agent = Customer.objects.create(
                user=user,
                email=email,
            )
            messages.success(request, 'Customer created successfully.')

    context = {
        'agent': agent,
        'customers': customers,
    }
    return render(request, 'dashboard/html/manage-customer.html', context)


def delete_agent_view(request, pk):
    agent = RentalAgent.objects.get(pk=pk)

    if request.method == 'POST':
        agent.delete()
        return redirect('manage-agent')

    return render(request, 'dashboard/html/manage-agent.html')


def delete_customer_view(request, pk):
    customer = Customer.objects.get(pk=pk)

    if request.method == 'POST':
        customer.delete()
        return redirect('manage-customer')

    return render(request, 'dashboard/html/manage-customer.html')


def manage_car_view(request):
    agent = RentalAgent.objects.get(user=request.user)
    cars = Car.objects.all()
    categories = Category.objects.all()

    search_by = request.GET.get('search_by', 'model')
    search_query = request.GET.get('search', '')
    available = request.GET.get('available', False)
    category = request.GET.get('category', False)

    # Filter based on the search parameters
    if search_by == 'model':
        cars = Car.objects.filter(model__icontains=search_query)
    else:
        # Handle other search criteria as needed
        pass

    # Additional filtering for availability and category
    if available:
        cars = cars.filter(available=True)

    if category:
        cars = cars.filter(category__name__icontains=category)

    if request.method == 'POST':
        if 'car_form_submit' in request.POST:
            car_form = CarForm(request.POST, request.FILES)
            if car_form.is_valid():
                car_form.save()
                messages.success(request, 'Car created successfully.')
        elif 'category_form_submit' in request.POST:
            category_form = CategoryForm(request.POST)
            if category_form.is_valid():
                category_form.save()
                messages.success(request, 'Category created successfully.')
    else:
        car_form = CarForm()
        category_form = CategoryForm()

    context = {
        'agent': agent,
        'cars': cars,
        'categories': categories,
    }

    return render(request, 'dashboard/html/manage-car.html', context)


def manage_category_view(request):
    agent = RentalAgent.objects.get(user=request.user)
    categories = Category.objects.all()

    if request.method == 'POST':
        category_form = CategoryForm(request.POST)
        if category_form.is_valid():
            category_form.save()
            messages.success(request, 'Category created successfully.')
    else:
        category_form = CategoryForm()

    context = {
        'agent': agent,
        'categories': categories,
    }

    return render(request, 'dashboard/html/manage-category.html', context)


def manage_rental_view(request):
    agent = RentalAgent.objects.get(user=request.user)
    customers = Customer.objects.all()
    cars = Car.objects.all()
    rental = Rental.objects.all()

    if request.method == 'POST':
        rental_form = RentalForm(request.POST)
        if rental_form.is_valid():
            rental_form.save()
            messages.success(request, 'Rental created successfully.')
    else:
        rental_form = RentalForm()

    context = {
        'agent': agent,
        'customers': customers,
        'cars': cars,
        'rental': rental,
    }

    return render(request, 'dashboard/html/manage-rental.html', context)


def manage_payment_view(request):
    agent = RentalAgent.objects.get(user=request.user)
    rental = Rental.objects.all()
    payments = Payment.objects.all()

    if request.method == 'POST':
        payment_form = PaymentForm(request.POST)
        if payment_form.is_valid():
            payment_form.save()
            messages.success(request, 'Payment created successfully.')
    else:
        payment_form = PaymentForm()

    context = {'agent': agent, 'rental': rental, 'payments': payments}

    return render(request, 'dashboard/html/manage-payment.html', context)


def edit_car_view(request, pk):
    car = Car.objects.get(id=pk)
    if request.method == 'POST':
        car_form = CarForm(request.POST, request.FILES, instance=car)
        if car_form.is_valid():
            car_form.save()
            messages.success(request, 'Car updated successfully.')
        return redirect('manage-car')
    else:
        car_form = CarForm(instance=car)

    return render(request, 'dashboard/html/manage-car.html')


def edit_category_view(request, pk):
    category = Category.objects.get(id=pk)
    if request.method == 'POST':
        category_form = CategoryForm(request.POST, instance=category)
        if category_form.is_valid():
            category_form.save()
            messages.success(request, 'Category updated successfully.')
        return redirect('manage-category')
    else:
        category_form = CategoryForm(instance=category)

    return render(request, 'dashboard/html/manage-category.html')


def edit_rental_view(request, pk):
    rental = Rental.objects.get(id=pk)

    if request.method == 'POST':
        rental_form = RentalForm(request.POST, instance=rental)
        if rental_form.is_valid():
            rental_form.save()
            messages.success(request, 'Rental updated successfully.')
        return redirect('manage-rental')
    else:
        rental_form = RentalForm(instance=rental)

    return render(request, 'dashboard/html/manage-rental.html')


def delete_car_view(request, pk):
    car = Car.objects.get(pk=pk)

    if request.method == 'POST':
        car.delete()
        return redirect('manage-car')

    return render(request, 'dashboard/html/manage-car.html')


def delete_category_view(request, pk):
    category = Category.objects.get(pk=pk)

    if request.method == 'POST':
        category.delete()
        return redirect('manage-category')

    return render(request, 'dashboard/html/manage-category.html')


def delete_rental_view(request, pk):
    rental = Rental.objects.get(pk=pk)

    if request.method == 'POST':
        rental.delete()
        return redirect('manage-rental')

    return render(request, 'dashboard/html/manage-rental.html')


def delete_payment_view(request, pk):
    payment = Payment.objects.get(pk=pk)

    if request.method == 'POST':
        payment.delete()
        return redirect('manage-payment')

    return render(request, 'dashboard/html/manage-payment.html')


def export_agents_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="rental_agents.csv"'

    writer = csv.writer(response)
    writer.writerow(['First Name', 'Last Name', 'Email', 'Phone'])

    rental_agents = RentalAgent.objects.all()
    for agent in rental_agents:
        writer.writerow([agent.first_name, agent.last_name, agent.email, agent.phone])

    return response


def export_agents_excel(request):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="rental_agents.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.append(['First Name', 'Last Name', 'Email', 'Phone'])

    rental_agents = RentalAgent.objects.all()
    for agent in rental_agents:
        ws.append([agent.first_name, agent.last_name, agent.email, agent.phone])

    wb.save(response)
    return response


def export_customers_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="customers.csv"'

    writer = csv.writer(response)
    writer.writerow(['First Name', 'Last Name', 'Email', 'Phone', 'Address'])

    customers = Customer.objects.all()
    for customer in customers:
        writer.writerow(
            [
                customer.first_name,
                customer.last_name,
                customer.email,
                customer.phone,
                customer.address,
            ]
        )

    return response


def export_customers_excel(request):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="customers.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.append(['First Name', 'Last Name', 'Email', 'Phone', 'Address'])

    customers = Customer.objects.all()
    for customer in customers:
        ws.append(
            [
                customer.first_name,
                customer.last_name,
                customer.email,
                customer.phone,
                customer.address,
            ]
        )

    wb.save(response)
    return response


def export_cars_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="cars.csv"'

    writer = csv.writer(response)
    writer.writerow(['Model', 'Make', 'Price', 'Category', 'Available'])

    cars = Car.objects.all()
    for car in cars:
        writer.writerow([car.model, car.make, car.price, car.category, car.available])

    return response


def export_cars_excel(request):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="cars.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.append(['Model', 'Make', 'Price', 'Category', 'Available'])

    cars = Car.objects.all()
    for car in cars:
        ws.append([car.model, car.make, car.price, car.category.category_name, car.available])

    wb.save(response)
    return response


def convert_to_vientiane_timezone(dt):
    if dt:
        local_tz = timezone.get_default_timezone()
        vientiane_offset = timezone.timedelta(hours=7)  # Asia/Vientiane is 7 hours ahead of UTC
        return dt - vientiane_offset
    else:
        return None


def convert_to_naive_datetime(dt):
    if dt:
        return dt.replace(tzinfo=None)
    else:
        return None


def export_rentals_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="rentals.csv"'

    writer = csv.writer(response)
    writer.writerow(['Customer', 'Car', 'Rental Start Date', 'Rental End Date', 'Total Cost'])

    rentals = Rental.objects.all()
    for rental in rentals:
        rental_start_date = convert_to_vientiane_timezone(rental.rental_start_date)
        rental_end_date = convert_to_vientiane_timezone(rental.rental_end_date)

        writer.writerow(
            [
                rental.customer,
                rental.car,
                rental_start_date,
                rental_end_date,
                rental.total_cost,
            ]
        )

    return response


def export_rentals_excel(request):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="rentals.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.append(['Customer', 'Car', 'Rental Start Date', 'Rental End Date', 'Total Cost'])

    rentals = Rental.objects.all()
    for rental in rentals:
        rental_start_date = convert_to_naive_datetime(
            convert_to_vientiane_timezone(rental.rental_start_date)
        )
        rental_end_date = convert_to_naive_datetime(
            convert_to_vientiane_timezone(rental.rental_end_date)
        )

        ws.append(
            [
                rental.customer.user.username,
                rental.car.model,
                rental_start_date,
                rental_end_date,
                rental.total_cost,
            ]
        )

    wb.save(response)
    return response


def table_view(request):
    return render(request, 'dashboard/html/table-basic.html')


def icon_material_view(request):
    agent = RentalAgent.objects.get(user=request.user)
    return render(request, 'dashboard/html/icon-material.html', {'agent': agent})


def map_google_view(request):
    agent = RentalAgent.objects.get(user=request.user)
    return render(request, 'dashboard/html/map-google.html', {'agent': agent})


def blank_view(request):
    agent = RentalAgent.objects.get(user=request.user)
    return render(request, 'dashboard/html/pages-blank.html', {'agent': agent})


def page_not_found_view(request):
    return render(request, 'dashboard/html/pages-error-404.html')
